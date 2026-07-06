"""Tests for the HTML-to-Excel conversion."""

import re
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from cycling.html_to_excel import convert_directory, sanitize_sheet_name


def test_sanitize_sheet_name_removes_forbidden_chars() -> None:
    used: set[str] = set()
    assert sanitize_sheet_name("Etappe 1/2 [A]", used) == "Etappe 1_2 _A_"


def test_sanitize_sheet_name_truncates_to_31() -> None:
    used: set[str] = set()
    name = sanitize_sheet_name("x" * 50, used)
    assert len(name) == 31


def test_sanitize_sheet_name_deduplicates() -> None:
    used: set[str] = set()
    first = sanitize_sheet_name("Ergebnis", used)
    second = sanitize_sheet_name("Ergebnis", used)
    assert first == "Ergebnis"
    assert second == "Ergebnis_2"


def _write_html_table(path: Path, df: pd.DataFrame) -> None:
    path.write_text(df.to_html(index=False), encoding="utf-8")


def test_convert_directory_one_sheet_per_file(tmp_path: Path) -> None:
    src = tmp_path / "html"
    src.mkdir()
    _write_html_table(
        src / "etappe1.html",
        pd.DataFrame({"Rang": [1, 2], "Fahrer": ["A", "B"]}),
    )
    _write_html_table(
        src / "etappe2.html",
        pd.DataFrame({"Rang": [1], "Fahrer": ["C"]}),
    )

    out = tmp_path / "out.xlsx"
    convert_directory(src, out)

    wb = load_workbook(out)
    assert set(wb.sheetnames) == {"etappe1", "etappe2"}
    assert wb["etappe1"]["B2"].value == "A"


def test_convert_directory_text_fallback(tmp_path: Path) -> None:
    src = tmp_path / "html"
    src.mkdir()
    (src / "notiz.html").write_text(
        "<html><body><p>Kein Tabelleninhalt hier</p></body></html>",
        encoding="utf-8",
    )

    out = tmp_path / "out.xlsx"
    convert_directory(src, out)

    wb = load_workbook(out)
    assert wb.sheetnames == ["notiz"]
    assert wb["notiz"]["A1"].value == "Inhalt"


def test_convert_directory_no_html_raises(tmp_path: Path) -> None:
    with pytest.raises(FileNotFoundError):
        convert_directory(tmp_path, tmp_path / "out.xlsx")


FIXTURES = Path(__file__).parent / "fixtures" / "html_results"


def test_fixtures_only_html_is_converted(tmp_path: Path) -> None:
    """Realdaten-Fixtures (anonymisiert): je HTML ein Sheet, PDFs ignoriert."""
    html_files = list(FIXTURES.glob("*.html"))
    pdf_files = list(FIXTURES.glob("*.pdf"))
    assert html_files, "Fixtures fehlen"
    assert pdf_files, "Fake-PDFs fehlen (Filter-Test)"

    out = tmp_path / "ergebnisse.xlsx"
    convert_directory(FIXTURES, out)

    wb = load_workbook(out)
    # Genau ein Sheet je HTML-Datei, kein Sheet für die PDFs.
    assert len(wb.sheetnames) == len(html_files)


def test_fixtures_are_anonymized() -> None:
    """Absicherung: die echten Event-Bezüge dürfen nie ins Repo gelangen."""
    blob = "".join(f.read_text(encoding="utf-8") for f in FIXTURES.glob("*.html"))
    for forbidden in ("Ebringen", "Badischer Radsportverband", "58. Radrenntag"):
        assert forbidden not in blob


def test_fixtures_drop_classes_below_u17() -> None:
    """Altersklassen unter U17 (Schüler U15/U13/U11) sind nicht enthalten."""
    names = [f.name for f in FIXTURES.glob("*.html")]
    for dropped in ("U15", "U13", "U11", "Schüler"):
        assert not any(dropped in n for n in names), f"{dropped} noch vorhanden"


def test_fixtures_race_grouping_has_varied_heat_counts() -> None:
    """Die Rennen enthalten je einmal 1, 2 und mehr als 2 Läufe."""
    counts: dict[str, int] = {}
    for f in FIXTURES.glob("*.html"):
        race = f.name.split(" - ")[0]  # "Rennen NN"
        counts[race] = counts.get(race, 0) + 1
    sizes = set(counts.values())
    assert 1 in sizes
    assert 2 in sizes
    assert any(s > 2 for s in sizes)


def test_fixtures_internal_race_number_matches_filename() -> None:
    """Die interne 'Rennen NN'-Referenz stimmt mit dem Dateinamen überein."""
    for f in FIXTURES.glob("*.html"):
        file_num = f.name.split(" ")[1]  # "01" aus "Rennen 01 - ..."
        match = re.search(r"Resultate von Rennen (\d+)", f.read_text(encoding="utf-8"))
        assert match is not None
        assert match.group(1) == file_num
