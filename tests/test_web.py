"""Tests für die Weboberfläche (Upload → Excel)."""

import io
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from cycling.web.main import app

client = TestClient(app)
FIXTURES = Path(__file__).parent / "fixtures" / "html_results"
XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def test_index_page_renders() -> None:
    resp = client.get("/")
    assert resp.status_code == 200
    assert "HTML-Ergebnisse in eine Excel-Datei" in resp.text
    assert "Kommissärs" in resp.text


def test_convert_uploads_produce_workbook() -> None:
    html_files = sorted(FIXTURES.glob("*.html"))[:3]
    uploads = [
        ("files", (f.name, f.read_bytes(), "text/html")) for f in html_files
    ]
    resp = client.post("/convert", files=uploads)

    assert resp.status_code == 200
    assert resp.headers["content-type"] == XLSX_MEDIA_TYPE
    assert "attachment" in resp.headers["content-disposition"]

    wb = load_workbook(io.BytesIO(resp.content))
    assert len(wb.sheetnames) == len(html_files)


def test_convert_ignores_non_html() -> None:
    html = sorted(FIXTURES.glob("*.html"))[0]
    pdf = sorted(FIXTURES.glob("*.pdf"))[0]
    uploads = [
        ("files", (html.name, html.read_bytes(), "text/html")),
        ("files", (pdf.name, pdf.read_bytes(), "application/pdf")),
    ]
    resp = client.post("/convert", files=uploads)

    assert resp.status_code == 200
    wb = load_workbook(io.BytesIO(resp.content))
    assert len(wb.sheetnames) == 1  # nur die HTML-Datei


def test_convert_without_html_returns_error() -> None:
    pdf = sorted(FIXTURES.glob("*.pdf"))[0]
    uploads = [("files", (pdf.name, pdf.read_bytes(), "application/pdf"))]
    resp = client.post("/convert", files=uploads)

    assert resp.status_code == 400
    assert "HTML-Datei" in resp.text
